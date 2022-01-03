import openpyxl
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws.cell(1,1).value = " "
ws.cell(1,2).value = "Test 1"
ws.cell(1,3).value = "Test 2"
ws.cell(1,4).value = "Test 3"
ws.cell(1,5).value = "Test 4"
ws.cell(1,6).value = "Test 5"
ws.cell(1,7).value = "Test 6"
ws.cell(1,8).value = "Test 7"

ws.cell(2,1).value = "Category"
ws.cell(2,2).value = "speakers"
ws.cell(2,3).value = "headphones"
ws.cell(2,4).value = "mice"
ws.cell(2,5).value = "mice"
ws.cell(2,6).value = "laptops"
ws.cell(2,7).value = "laptops"
ws.cell(2,8).value = "tablets"

ws.cell(3,1).value = "Product ID"
ws.cell(3,2).value = "19"
ws.cell(3,3).value = "12"
ws.cell(3,4).value = "28"
ws.cell(3,5).value = "30"
ws.cell(3,6).value = "9"
ws.cell(3,7).value = "7"
ws.cell(3,8).value = "18"

ws.cell(4,1).value = "Quantity"
ws.cell(4,2).value = "2"
ws.cell(4,3).value = "2"
ws.cell(4,4).value = " "
ws.cell(4,5).value = "1"
ws.cell(4,6).value = "2"
ws.cell(4,7).value = "3"
ws.cell(4,8).value = " "

ws.cell(5,1).value = "Color"
ws.cell(5,2).value = "RED"
ws.cell(5,3).value = "PURPLE"
ws.cell(5,4).value = "GREEN"
ws.cell(5,5).value = " "
ws.cell(5,6).value = "BLUE"
ws.cell(5,7).value = " "
ws.cell(5,8).value = "yellow"

wb.save('openpyxl.xlsx')